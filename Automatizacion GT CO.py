import re
import openpyxl


def encontrar_decimales(expresion, decimales):
    # Reemplazar "DD" por los números decimales encontrados
    try:
        for decimal in decimales:
         expresion = expresion.replace("DD", decimal, 1)
    
    except:
        expresion = expresion

    return expresion

def evaluate_expression(expression, dictionary):
    patron = r'\d+\.\d+'
    
    # Buscar todos los números decimales en la expresión
    decimales = re.findall(patron, expression)
    
    # Reemplazar los decimales por "DD" en la expresión
    expresion_modificada = re.sub(patron, "DD", expression)
    

    numeros_y_signos = re.findall(r'[+\-*/()]|\d+(?:\.\d+)?|DD', expresion_modificada)

    n_y_s = ''
    
    for numero in numeros_y_signos:
        for key, value in dictionary.items():
            if str(key) == str(numero):
                numero = numero.replace(str(key), str(value))
                numero = '('+numero+')'
                break
    
    
    
        n_y_s = n_y_s + numero
        
            
    expresion_modificada = n_y_s      
    
    
    
    expression = encontrar_decimales(expresion_modificada,decimales)
       


    return eval(expression)


# Abre el archivo de Excel
workbook = openpyxl.load_workbook("C:\\Users\\Walter Montiel\\Desktop\\GT CO\\Nuevos\\Automatizacion Kaika.xlsx")

# Obtiene las hojas de cálculo
sheet1 = workbook['Balance']
sheet2 = workbook['Renglon']
sheet3 = workbook['Formulas']

# Diccionario para almacenar los grupos y sus sumas
grupos = {}

# Calcula la suma de cada grupo en la pestaña 2
for row in sheet2.iter_rows(min_row=2,values_only=True):

    grupo = row[2]
    monto = row[3]
    
    
    
    if grupo in grupos:
        grupos[grupo] += monto
    else:
        grupos[grupo] = monto



# Procesa las fórmulas en la pestaña 3 y guarda los resultados en la pestaña 2
for row in sheet3.iter_rows(min_row=2, values_only=True):
    renglon = row[0]
    formula = row[1]

  
    result = evaluate_expression(str(formula).replace(',','.'), grupos.copy())
    print("Result:", result)