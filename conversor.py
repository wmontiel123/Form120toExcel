import PyPDF2
import re
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog
from tkinter.filedialog import askdirectory
import os


def convertir_fecha(fecha):
    fecha_convertida = fecha
    try:
        partes = fecha.split()
        mes = partes[0].rjust(2, '0')
        anio = ''.join(partes[2:])
        fecha_convertida = f"{mes}-{anio}"
    except:
        pass
    
    return fecha_convertida

def get_pdf_files(folder_path):
    pdf_files = []
    try:
        for file_name in os.listdir(folder_path):
            if file_name.endswith('.pdf'):
                pdf_files.append(file_name)
    except:
        pass
    return pdf_files

def leer_pdf(ruta):
    lineas = []
    try:
        with open(ruta, 'rb') as archivo:
            lector_pdf = PyPDF2.PdfReader(archivo)
            for pagina in lector_pdf.pages:
                contenido = pagina.extract_text()
                lineas.extend(contenido.split('\n'))
    except:
        pass

    return lineas

def extract_text(linea, inicio, fin1, fin2):
   # Find the position of the inicio string
    inicio_pos = linea.find(inicio)
    if inicio_pos == -1:
        return None
    fin_pos = len(linea)+2
    if fin1 != "" and fin2 !="":
        fin_pos = linea.find(fin1, inicio_pos)

    # Find the position of the fin2 string
    if fin_pos == -1:
        fin_pos = linea.find(fin2, inicio_pos)+1
        inicio_pos = inicio_pos
        
    if fin_pos == -1:
        return None 



    # Extract the text between the inicio_pos and fin_pos positions
    texto_extraido = linea[(inicio_pos) + len(inicio):fin_pos]
    return texto_extraido.strip()  # Remove leading/trailing whitespaces


# Ruta del archivo PDF
# Create a Tkinter root window
root = tk.Tk()
root.withdraw()  # Hide the root window

# Ask user to select a folder using a dialog box
folder_path = askdirectory(title="Select Folder")
if not folder_path:
    print("No folder selected. Exiting...")
    exit()

pdf_files = get_pdf_files(folder_path)

columna1 = 1
columna2 = 2

# Crear un nuevo libro de Excel y seleccionar la hoja activa
wb = Workbook()
sheet = wb.active
for file_path in pdf_files:
    # Print the selected file path
   
    ruta_pdf = folder_path+'\\'+ file_path  
    Valores_Obtenidos = []
    values = [
        ("vados con tasa del 10%10", " 22 ", " 022 "),
        (" 22 ", "", ""),
        (" 022 ", "", ""),
        ("con tasa del 5%150", " 156 ", "0156 "),
        (" 156 ", "", ""),
        (" 0156 ", "", ""),
        ("gravados con tasa del 5%151", " 157 ", "0157 "),
        (" 157 ", "", ""),
        (" 0157 ", "", ""),
        ("no alcanzados por el Impuesto12", "", ""),
        ("industrializacion152 ", "", ""),
        ("bienes153 ", "", ""),
        ("bienes 14", "", ""),
        ("tasa del 10%15", " 23 ", "023 "),
        (" 23 ", "", ""),
        (" 023 ", "", ""),
        ("proceso de elaboracion o industrializacion154", " 158 ", "0158 "),
        (" 158 ", "", ""),
        (" 0158 ", "", ""),
        ("servicios155 ", " 159 ", "0159 "),
        (" 0159 ", "", ""),
        (" 159 ", "", ""),
        ("Impuesto17 ", "", ""),
        ("Inc. a+h)18", " 21 ", "021 "),
        (" 21 ", " 24 ", "024 "),
        (" 021 ", " 24 ", "024 "),
        (" 24 ", "", ""),
        (" 024 ", "", ""),
        ("natural160", "", ""),
        ("interno161", "", ""),
        ("no alcanzados 26", "", ""),
        ("(Inc. a+b+c) 27", "", ""),
        ("industrializacion162", "", ""),
        ("bienes 163", "", ""),
        ("bienes 29", "", ""),
        ("(Inc. e+f+g) 30", "", ""),
        ("(Inc. d+h) 31", "", ""),
        ("interno32 ", " 35 ", "035 "),
        (" 35 ", " 38 ", "038 "),
        (" 035 ", " 38 ", "038 "),
        (" 038 ", "", ""),
        (" 38 ", "", ""),
        ("alcanzadas33", " 36 ", "036 "),
        (" 36 ", " 39 ", "039 "),
        (" 036 ", " 39 ", "039 "),
        (" 39 ", "", ""),
        (" 039 ", "", ""),
        ("Rubro 2 Inc. a+b/ Inc. d)164", "", ""),
        ("calculo)165", "", ""),
        ("incobrables34", " 37 ", "037 "),
        (" 37 ", " 42 ", "042 "),
        (" 037 ", " 42 ", "042 "),
        (" 42 ", "", ""),
        (" 042 ", "", ""),
        ("a+c+d+e)43", "", ""),
        ("Inc. l) 44", "", ""),
        (" Inc. f) 45", "", ""),
        ("anterior)46", "", ""),
        ("Inc. c 166", "", ""),
        ("d167 ", "", ""),
        ("e47", "", ""),
        ("Inc. c 48", "", ""),
        (" Exportador)49", "", ""),
        ("(No trasladable)168", "", ""),
        ("i) 50", "", ""),
        ("Inc. j) 55", "", ""),
        ("anterior)51", "", ""),
        ("recibidas52", "", ""),
        ("gravadas 169", "", ""),
        ("vencimiento 56", "", ""),
        (" a+e) 53", " 57 ", "057 "),
        (" 57 ", "", ""),
        (" 057 ", "", ""),
        ("Rubro 454", "", ""),
        ("Col. I)58", "", ""),
        ("Impuesto59", " 65 ", "065 "),
        (" 65 ", "", ""),
        (" 065 ", "", ""),
        ("Impuesto60", " 66 ", "066 "),
        (" 66 ", "", ""),
        (" 066 ", "", ""),
        ("exportaciones61", "", ""),
        ("Impuesto62", "", ""),
        ("Impuesto63", "", ""),
        ("IRP 64 ", "", ""),
        ("IRP170 ", "", ""),
        #Anexo Exportador
        #("PERIODO ATRIBUIDAS A EXPORTACION", "", ""),
        ("industrializacion171", " 181 ", "0181 "),
        ("servicios atribuido directamente a exportacion172", " 177 ", "0177 "),
        ("anticipadamente173", " 183 ", "0183 "),
        ("exportaciones174", " 178 ", "0178 "),
        ("exportaciones175", " 179 ", "0179 "),
        ("exportaciones176", " 180 ", "0180 "),
        (" 177 ", " 182 ", "0182 "),
        (" 0177 ", " 182 ", "0182 "),
        (" 178 ", " 184 ", "0184 "),
        (" 0178 ", " 184 ", "0184 "),
        (" 179 ", " 185 ", "0185 "),
        (" 0179 ", " 185 ", "0185 "),
        (" 180 ", " 186 ", "0186 "),
        (" 0180 ", " 186 ", "0186 "),
        (" 181 ", "", ""),
        (" 0181 ", "", ""),
        (" 182 ", "", ""),
        (" 0182 ", "", ""),
        (" 183 ", "", ""),
        (" 0183 ", "", ""),
        (" 184 ", "", ""),
        (" 0184 ", "", ""),
        (" 185 ", "", ""),
        (" 0185 ", "", ""),
        (" 186 ", "", ""),
        (" 0186 ", "", ""),
        ("servicios atribuido directamente a fletes de exportacion187 ", " 195 ", "0195 "),
        ("exportacion188", " 196 ", "0196 "),
        ("exoneradas o no alcanzadas y fletes de exportacion189", " 197 ", "0197 "),
        ("exportacion190", " 198 ", "0198 "),
        ("exportacion191", " 199 ", "0199 "),
        ("en el mercado interno y fletes de exportacion192", " 200 ", "0200 "),
        ("exoneradas o no alcanzadas y fletes de exportacion193", " 201 ", "0203 "),
        ("de exportacion194", " 202 ", "0202 "),
        ("de exportacion194", " 202 ", "0202 "),
        (" 195 ", " 203 ", "0203 "),
        (" 0195 ", " 203 ", "0203 "),
        (" 196 ", " 204 ", "0204 "),
        (" 0196 ", " 204 ", "0204 "),
        (" 197 ", " 205 ", "0205 "),
        (" 0197 ", " 205 ", "0205 "),
        (" 198 ", " 206 ", "0206 "),
        (" 0198 ", " 206 ", "0206 "),
        (" 199 ", " 207 ", "0207 "),
        (" 0199 ", " 207 ", "0207 "),
        (" 200 ", " 208 ", "0208 "),
        (" 0200 ", " 208 ", "0208 "),
        (" 201 ", " 209 ", "0209 "),
        (" 0201 ", " 209 ", "0209 "),
        (" 202 ", " 210 ", "0210 "),
        (" 0202 ", " 210 ", "0210 "),
        (" 203 ", "", ""),
        (" 0203 ", "", ""),
        (" 204 ", "", ""),
        (" 0204 ", "", ""),
        (" 205 ", "", ""),
        (" 0205 ", "", ""),
        (" 206 ", "", ""),
        (" 0206 ", "", ""),
        (" 207 ", "", ""),
        (" 0207 ", "", ""),
        (" 208 ", "", ""),
        (" 0208 ", "", ""),
        (" 209 ", "", ""),
        (" 0209 ", "", ""),
        (" 210 ", "", ""),
        (" 0210 ", "", ""),
        ("(Proviene de la casilla 182)211", "", ""),
        ("(Proviene de la casilla 203)212", "", ""),
        ("IVA Credito atribuido proporcionalmente a exportacion (Proviene de la casilla 218 de la Hoja de calculo) 213", "", ""),
        ("(Proviene de la suma de las casillas 211+212+213)214", "", ""),
        ("solicitudes de devolucion del IVA Credito del Exportador presentadas en el periodo anterior al que se liquida)215", "", ""),
        ("2/2Importe del IVA Credito por exportacion aplicado al mercado interno 148", "", ""),
        ("(Casillas 214+215-148)149", "", "")
        
        
    ]



    # Llamada a la función para leer el PDF y guardar las líneas en una lista
    lineas_pdf = leer_pdf(ruta_pdf)
    
    #Obtener datos de Cabecera
    contribuyente = ''
    control = ''
    fecha = ''
    Orden = ''
    Mes_Año = ''
    Tipo_declaracion = ''
    Normalizada = ''
    Anexo_Exportador = 'NO'
    
    for i in range(len(lineas_pdf)-1):
        if lineas_pdf[i].startswith("Mes Año"):
            Mes_Año = lineas_pdf[i+1]
            break
    
    
    for line in lineas_pdf:
        if "DECLARACIÓN JURADA NORMALIZADA" in line:
            if "DECLARACIÓN JURADA NORMALIZADA" in line:
                Normalizada = 'DECLARACIÓN JURADA NORMALIZADA'
            else:
                Normalizada = 'DECLARACIÓN JURADA ORIGINAL'
                
            break
        
    numero_de_linea = 0
    linea_contribuyente = 0
    for line in lineas_pdf:
        #print(line)
        numero_de_linea = numero_de_linea + 1
        if line.startswith("COMPRAS LOCALES E IMPORTACIONES DEL"):
            Anexo_Exportador = re.search(r'COMPRAS LOCALES E IMPORTACIONES DEL\s+(\d+)', line)
            if Anexo_Exportador:
                Anexo_Exportador = 'SI'
            
        
        if line.startswith("Número de Orden"):
            Orden = re.search(r'Número de Orden\s+(\d+)', line)
            linea_contribuyente = numero_de_linea
            if Orden:
                Orden = Orden.group(1)
                
        
            
        if line.startswith("Formulario"):
            # Extract Contribuyente using regular expression
            contribuyente = re.search(r'Contribuyente:\s+(\d+)', line)
            if contribuyente:
                contribuyente = contribuyente.group(1)

            # Extract Control using regular expression
            control = re.search(r'Control:\s+(\w+)', line)
            if control:
                control = control.group(1)

            # Extract Fecha using regular expression
            fecha = re.search(r'Fecha:\s+([\d/]+\s+[\d:]+)', line)
            if fecha:
                fecha = fecha.group(1)
           
        
        if line.startswith("02Declaración Jurada Rectificativa 03"):
            # Extract Contribuyente using regular expression
            Tipo_declaracion = re.search(r'02Declaración Jurada Rectificativa 03\s+(\d+)', line)
            if Tipo_declaracion:
                Tipo_declaracion = Tipo_declaracion.group(1)

        
        
        
    
    # Imprimir las líneas del PDF
    
    #print(Tipo_declaracion)
    DV = lineas_pdf[linea_contribuyente]
    for linea in lineas_pdf:
        
        for inicio, fin1, fin2 in values:
            texto_extraido = extract_text(linea, inicio, fin1, fin2)
            if texto_extraido:
                
                #print(linea)
                
                #print(inicio)
                
                #print(texto_extraido)
                #print(linea)
                
                Valores_Obtenidos.append(texto_extraido)
                
                

    

    # Datos que deseas copiar en la columna B
    datos = []
    if Anexo_Exportador == 'NO':
        datos = ["10",
            "22",
            "150",
            "156",
            "151",
            "157",
            "12",
            "152",
            "153",
            "14",
            "15",
            "23",
            "154",
            "158",
            "155",
            "159",
            "17",
            "18",
            "21",
            "24",
            "160",
            "161",
            "26",
            "27",
            "162",
            "163",
            "29",
            "30",
            "31",
            "32",
            "35",
            "38",
            "33",
            "36",
            "39",
            "164",
            "165",
            "34",
            "37",
            "42",
            "43",
            "44",
            "45",
            "46",
            "166",
            "167",
            "47",
            "48",
            "49",
            "168",
            "50",
            "55",
            "51",
            "52",
            "169",
            "56",
            "53",
            "57",
            "54",
            "58",
            "59",
            "65",
            "60",
            "66",
            "61",
            "62",
            "63",
            "64",
            "170"
        ]
    else:
       datos = [   
            "171",
            "181",
            "172",
            "177",
            "182",
            "173",
            "183",
            "174",
            "178",
            "184",
            "175",
            "179",
            "185",
            "176",
            "180",
            "186",
            "187",
            "195",
            "203",
            "188",
            "196",
            "204",
            "189",
            "197",
            "205",
            "190",
            "198",
            "206",
            "191",
            "199",
            "207",
            "192",
            "200",
            "208",
            "193",
            "201",
            "209",
            "194",
            "202",
            "210",
            "211",
            "212",
            "213",
            "214",
            "215",
            "148",
            "149"
            ]

    # Copiar los datos en el excel
    
    sheet.cell(row = 1, column = columna1).value =  Normalizada
    sheet.cell(row = 2, column = columna1).value = 'Contribuyente'
    sheet.cell(row = 3, column = columna1).value = 'Control'
    sheet.cell(row = 4, column = columna1).value = 'Fecha/Hora'
    sheet.cell(row = 5, column = columna1).value = 'Numero de Orden'
    sheet.cell(row = 6, column = columna1).value = 'Mes/Año'
    sheet.cell(row = 7, column = columna1).value = 'Tipo Declaracion'
    
    sheet.cell(row = 1, column = columna2).value = ''
    sheet.cell(row = 2, column = columna2).value = contribuyente + '-' + DV
    sheet.cell(row = 3, column = columna2).value = control
    sheet.cell(row = 4, column = columna2).value = convertir_fecha(fecha)
    sheet.cell(row = 5, column = columna2).value = Orden
    sheet.cell(row = 6, column = columna2).value = Mes_Año
    
    if Tipo_declaracion == '0':
        sheet.cell(row = 7, column = columna2).value = 'ORIGINAL'
    else:
        sheet.cell(row = 7, column = columna2).value = 'RECTIFICATIVA'
    
    for i, dato in enumerate(datos, start=8):
        #print(dato)
        sheet.cell(row=i, column = columna1).value = dato
        
    for i, dato in enumerate(Valores_Obtenidos, start=8):
        #print(dato)
        sheet.cell(row=i,column = columna2).value = dato
    
    columna1 = columna1+2
    columna2 = columna2+2




# Guardar el archivo de Excel con los datos copiados
# Create a Tkinter root window
root = tk.Tk()
root.withdraw()  # Hide the root window

# Open the file dialog to select a file path for saving
file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=(("xlsx file", "*.xlsx"),("All Files", "*.*") ))

# Print the selected file path
print("Selected File Path for Saving:", file_path)

nombre_archivo = file_path

wb.save(nombre_archivo)
            
    

