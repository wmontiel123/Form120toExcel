

import openpyxl
import re

# Abre el archivo de Excel
workbook = openpyxl.load_workbook("C:\\Users\\Walter Montiel\\Desktop\\GT PY\\test.xlsx")

# Obtiene las hojas de cálculo
sheet1 = workbook['Pestaña 1']
sheet2 = workbook['Pestaña 2']
sheet3 = workbook['Pestaña 3']

# Diccionario para almacenar los grupos y sus sumas
grupos = {}

# Calcula la suma de cada grupo en la pestaña 1
for row in sheet1.iter_rows(min_row=2, values_only=True):

    grupo = row[0]
    monto = row[3]
    
    
    
    if grupo in grupos:
        grupos[grupo] += monto
    else:
        grupos[grupo] = monto



# Procesa las fórmulas en la pestaña 3 y guarda los resultados en la pestaña 2
for row in sheet3.iter_rows(min_row=2, values_only=True):
    renglon = row[0]
    formula = row[1]
    
    print(renglon,formula)
    
    # Evalúa las operaciones en la fórmula
    operaciones = re.findall(r'\d+(\.\d+)?\s*[-+\/*]\s*(?:\d+(\.\d+)?|\$\d+\$)', formula)
    resultado = None
    print(operaciones)
    
    for operacion in operaciones:
        if operacion.startswith('$') and operacion.endswith('$'):
            # Procesa valores entre "$" como números enteros
            valor = int(operacion.strip('$'))
        else:
            print(operacion)
            # Procesa valores numéricos
            valor = float(operacion)
        
        if resultado is None:
            resultado = valor
        else:
            # Realiza la operación correspondiente
            if operacion == '+':
                resultado += valor
            elif operacion == '-':
                resultado -= valor
            elif operacion == '*':
                resultado *= valor
            elif operacion == '/':
                resultado /= valor
    
    # Guarda el resultado en la pestaña 2
    sheet2.cell(row=renglon, column=2).value = resultado

# Guarda los cambios en el archivo
workbook.save("C:\\Users\\Walter Montiel\\Desktop\\GT PY\\test.xlsx")
